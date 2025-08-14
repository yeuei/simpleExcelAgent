import uuid
import logging

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from .exceptions import DataError

logger = logging.getLogger(__name__)

def create_excel_table(
    filepath: str,
    sheet_name: str,
    data_range: str,
    table_name: str | None = None,
    table_style: str = "TableStyleMedium9"
) -> dict:
    """Creates a native Excel table for the given data range.
    
    Args:
        filepath: Path to the Excel file.
        sheet_name: Name of the worksheet.
        data_range: The cell range for the table (e.g., "A1:D5").
        table_name: A unique name for the table. If not provided, a unique name is generated.
        table_style: The visual style to apply to the table.
        
    Returns:
        A dictionary with a success message and table details.
    """
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            raise DataError(f"Sheet '{sheet_name}' not found.")
            
        ws = wb[sheet_name]

        # If no table name is provided, generate a unique one
        if not table_name:
            table_name = f"Table_{uuid.uuid4().hex[:8]}"

        # Check if table name already exists
        if table_name in ws.parent.defined_names:
            raise DataError(f"Table name '{table_name}' already exists.")

        # Create the table
        table = Table(displayName=table_name, ref=data_range)
        
        # Apply style
        style = TableStyleInfo(
            name=table_style, 
            showFirstColumn=False,
            showLastColumn=False, 
            showRowStripes=True, 
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        
        ws.add_table(table)
        
        wb.save(filepath)
        
        return {
            "message": f"Successfully created table '{table_name}' in sheet '{sheet_name}'.",
            "table_name": table_name,
            "range": data_range
        }

    except Exception as e:
        logger.error(f"Failed to create table: {e}")
        raise DataError(str(e)) 